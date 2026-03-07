from __future__ import annotations

import argparse
import os
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

import gspread
import openpyxl
from google.oauth2.service_account import Credentials

try:
    import tomllib
except ModuleNotFoundError:  # pragma: no cover - fallback for older Python
    import tomli as tomllib  # type: ignore

from crm.core.config import SHEETS_SCHEMA

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]
DEFAULT_LEGACY_FILENAME = "Fechamento Motorista cegonha.xlsx"
DEFAULT_PREVIEW_FILENAME = "preview_migracao_cegonhas.xlsx"


class MigrationError(RuntimeError):
    """Erro de execucao da migracao."""


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_token(value: Any) -> str:
    text = normalize_text(value).lower()
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def parse_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = normalize_text(value)
    if not text:
        return None

    text = text.replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def safe_int(value: Any) -> int:
    text = normalize_text(value).replace(",", ".")
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def canonical_id(value: Any) -> str:
    text = normalize_text(value).replace(",", ".")
    if not text:
        return ""
    try:
        as_float = float(text)
        if as_float.is_integer():
            return str(int(as_float))
    except ValueError:
        pass
    return text


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = normalize_text(value)
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def fmt_date(value: date | None) -> str:
    return value.strftime("%Y-%m-%d") if value else ""


def month_comp(value: date | None) -> str:
    return f"{value.year}-{value.month:02d}" if value else ""


def resolve_secrets_path(explicit: str | None) -> Path | None:
    candidates: list[Path] = []
    if explicit:
        candidates.append(Path(explicit).expanduser())
    candidates.extend(
        [
            Path(".streamlit/secrets.toml"),
            Path(__file__).resolve().parent / ".streamlit" / "secrets.toml",
        ]
    )

    seen: set[Path] = set()
    for candidate in candidates:
        resolved = candidate.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        if resolved.exists():
            return resolved
    return None


def load_secrets(path: Path | None) -> dict[str, Any]:
    if not path:
        return {}
    with path.open("rb") as fp:
        data = tomllib.load(fp)
    return data if isinstance(data, dict) else {}


def get_credentials(args: argparse.Namespace, secrets: dict[str, Any]) -> Credentials:
    if args.credentials:
        creds_path = Path(args.credentials).expanduser()
        if not creds_path.exists():
            raise MigrationError(f"Arquivo de credencial nao encontrado: {creds_path}")
        return Credentials.from_service_account_file(str(creds_path), scopes=SCOPES)

    env_creds = os.getenv("CRM_CEGONHAS_CREDS") or os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
    if env_creds:
        creds_path = Path(env_creds).expanduser()
        if creds_path.exists():
            return Credentials.from_service_account_file(str(creds_path), scopes=SCOPES)

    service_account_info = secrets.get("gcp_service_account")
    if isinstance(service_account_info, dict) and service_account_info.get("client_email"):
        return Credentials.from_service_account_info(dict(service_account_info), scopes=SCOPES)

    for fallback_name in ("credentials.json", "service_account.json"):
        fallback_path = Path(fallback_name)
        if fallback_path.exists():
            return Credentials.from_service_account_file(str(fallback_path), scopes=SCOPES)

    raise MigrationError(
        "Credenciais nao encontradas. Use --credentials, variavel de ambiente "
        "CRM_CEGONHAS_CREDS/GOOGLE_APPLICATION_CREDENTIALS ou .streamlit/secrets.toml."
    )


def resolve_target(args: argparse.Namespace, secrets: dict[str, Any]) -> tuple[str, str]:
    spreadsheet_id = normalize_text(
        args.spreadsheet_id or os.getenv("CRM_SPREADSHEET_ID") or secrets.get("spreadsheet_id")
    )
    spreadsheet_name = normalize_text(
        args.spreadsheet_name or os.getenv("CRM_SPREADSHEET_NAME") or secrets.get("spreadsheet_name")
    )

    if not spreadsheet_id and not spreadsheet_name:
        raise MigrationError(
            "Planilha de destino nao definida. Informe --spreadsheet-id ou --spreadsheet-name, "
            "ou configure em .streamlit/secrets.toml."
        )
    return spreadsheet_id, spreadsheet_name


def resolve_legacy_path(explicit: str | None) -> Path:
    candidates: list[Path] = []
    env_legacy = os.getenv("CRM_LEGACY_XLSX")

    if explicit:
        candidates.append(Path(explicit).expanduser())
    if env_legacy:
        candidates.append(Path(env_legacy).expanduser())

    desktop_candidates = [Path.home() / "Desktop"]
    userprofile = os.getenv("USERPROFILE")
    if userprofile:
        desktop_candidates.append(Path(userprofile) / "OneDrive" / "Desktop")
    onedrive = os.getenv("ONEDRIVE")
    if onedrive:
        desktop_candidates.append(Path(onedrive) / "Desktop")

    desktop_folders = []
    for desktop in desktop_candidates:
        desktop_folders.extend(
            [
                desktop,
                desktop / "migracao_cegonhas",
                desktop / "migração_cegonhas",
            ]
        )

    candidates.extend(
        [
            Path(DEFAULT_LEGACY_FILENAME),
            Path(__file__).resolve().parent / DEFAULT_LEGACY_FILENAME,
            Path("..") / "migracao_cegonhas" / DEFAULT_LEGACY_FILENAME,
            Path("..") / "migração_cegonhas" / DEFAULT_LEGACY_FILENAME,
        ]
    )
    candidates.extend(folder / DEFAULT_LEGACY_FILENAME for folder in desktop_folders)

    inspected: list[str] = []
    seen: set[Path] = set()
    for candidate in candidates:
        resolved = candidate.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        inspected.append(str(resolved))
        if not resolved.exists():
            continue
        if resolved.stat().st_size <= 0:
            raise MigrationError(
                f"Arquivo legado encontrado, mas vazio (0 bytes): {resolved}. "
                "Substitua por uma copia valida."
            )
        return resolved

    inspected_list = "\n - ".join(inspected)
    raise MigrationError(
        "Nao achei o arquivo legado .xlsx. Caminhos verificados:\n - " + inspected_list
    )


def get_header_and_rows(worksheet: gspread.Worksheet) -> tuple[list[str], list[dict[str, str]]]:
    values = worksheet.get_all_values()
    if not values:
        raise MigrationError(
            f'Aba "{worksheet.title}" vazia. Abra o app uma vez para inicializar cabecalhos.'
        )

    header = [normalize_text(col) for col in values[0]]
    rows: list[dict[str, str]] = []
    for raw_row in values[1:]:
        row_dict = {
            column: normalize_text(raw_row[idx]) if idx < len(raw_row) else ""
            for idx, column in enumerate(header)
        }
        rows.append(row_dict)
    return header, rows


def ensure_required_columns(tab_name: str, header: list[str]) -> None:
    required = SHEETS_SCHEMA[tab_name]
    missing = [column for column in required if column not in header]
    if missing:
        raise MigrationError(
            f'Aba "{tab_name}" sem colunas obrigatorias: {missing}. '
            "Nao vou alterar schema automaticamente para evitar risco."
        )


def row_from_payload(payload: dict[str, Any], header: list[str]) -> list[Any]:
    return [payload.get(column, "") for column in header]


def extract_frota_id(sheet_name: str, row1: Any) -> int | None:
    match = re.search(r"(\d{3,4})", sheet_name)
    if match:
        return int(match.group(1))
    if row1:
        match = re.search(r"(\d{3,4})", str(row1))
        if match:
            return int(match.group(1))
    return None


def extract_driver(row1: Any) -> str:
    if not row1:
        return ""
    text = str(row1)
    text = re.sub(r"FROTA.*", "", text, flags=re.IGNORECASE).strip(" -/")
    text = re.sub(r"\s+", " ", text)
    return text.title() if text else ""


def find_header_cols(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, int, int, int]:
    max_scan_row = min(ws.max_row, 20)
    max_scan_col = min(ws.max_column, 12)

    for row_idx in range(1, max_scan_row + 1):
        tokens = [normalize_token(ws.cell(row_idx, col_idx).value) for col_idx in range(1, max_scan_col + 1)]
        if "data" not in tokens:
            continue

        local_col = None
        for key in ("local", "destino"):
            if key in tokens:
                local_col = tokens.index(key) + 1
                break

        value_col = None
        for key in ("valor", "frete"):
            if key in tokens:
                value_col = tokens.index(key) + 1
                break

        if local_col and value_col:
            return row_idx, tokens.index("data") + 1, local_col, value_col

    raise ValueError(f"Nao encontrei cabecalho Data/Local/Valor na aba {ws.title}")


def parse_legacy_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, Any]:
    header_row, date_col, local_col, value_col = find_header_cols(ws)
    row1 = ws.cell(1, 1).value or ws.cell(1, 2).value
    frota_id = extract_frota_id(ws.title, row1)
    motorista = extract_driver(row1)

    viagens: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_date = ws.cell(row_idx, date_col).value
        raw_destino = ws.cell(row_idx, local_col).value
        raw_frete = ws.cell(row_idx, value_col).value

        destino = normalize_text(raw_destino)
        if "TOTAL" in destino.upper():
            continue

        data_viagem = parse_date(raw_date)
        frete = parse_float(raw_frete)
        if data_viagem and destino and frete is not None:
            viagens.append(
                {
                    "data": data_viagem,
                    "destino": destino,
                    "frete_total": round(frete, 2),
                }
            )

    abastecimentos: list[dict[str, Any]] = []
    fuel_header_row = None
    for row_idx in range(1, ws.max_row + 1):
        first_col = ws.cell(row_idx, 1).value
        if isinstance(first_col, str) and "ABAST" in first_col.upper():
            fuel_header_row = row_idx
            break

    if fuel_header_row:
        for row_idx in range(fuel_header_row + 1, ws.max_row + 1):
            row_values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, 9)]
            if any(isinstance(item, str) and "TOTAL" in item.upper() for item in row_values):
                break

            valor = parse_float(ws.cell(row_idx, 3).value)
            data_item = parse_date(ws.cell(row_idx, 1).value)
            if valor is not None and valor > 0:
                abastecimentos.append({"data": data_item, "valor": round(valor, 2)})

    outros_gastos_total: float | None = None
    for row_idx in range(1, ws.max_row + 1):
        col_e = normalize_token(ws.cell(row_idx, 5).value)
        col_f = parse_float(ws.cell(row_idx, 6).value)
        if col_e == "total" and col_f is not None and col_f > 0:
            outros_gastos_total = round(col_f, 2)

    return {
        "sheet_name": ws.title,
        "frota_id": frota_id,
        "motorista": motorista,
        "viagens": viagens,
        "abastecimentos": abastecimentos,
        "outros_gastos_total": outros_gastos_total,
    }


def date_key(value: Any) -> str:
    parsed = parse_date(value)
    return fmt_date(parsed) if parsed else normalize_text(value)


def amount_key(value: Any) -> float | None:
    parsed = parse_float(value)
    return round(parsed, 2) if parsed is not None else None


def build_preview(
    preview_path: Path,
    summary_rows: list[list[Any]],
    frotas_header: list[str],
    viagens_header: list[str],
    despesas_header: list[str],
    frotas_rows: list[list[Any]],
    viagens_rows: list[list[Any]],
    despesas_rows: list[list[Any]],
) -> None:
    workbook = openpyxl.Workbook()

    ws_summary = workbook.active
    ws_summary.title = "resumo"
    ws_summary.append(
        [
            "sheet",
            "frota_id",
            "motorista",
            "viagens_encontradas",
            "abastecimentos_encontrados",
            "gastos_agregados_total",
            "viagens_importadas",
            "despesas_importadas",
        ]
    )
    for row in summary_rows:
        ws_summary.append(row)

    ws_frotas = workbook.create_sheet("frotas_append")
    ws_frotas.append(frotas_header)
    for row in frotas_rows:
        ws_frotas.append(row)

    ws_viagens = workbook.create_sheet("viagens_append")
    ws_viagens.append(viagens_header)
    for row in viagens_rows:
        ws_viagens.append(row)

    ws_despesas = workbook.create_sheet("despesas_append")
    ws_despesas.append(despesas_header)
    for row in despesas_rows:
        ws_despesas.append(row)

    preview_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(preview_path)


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Migra historico legado para o Google Sheets do CRM Cegonhas sem alterar schema, "
            "com dry-run padrao."
        )
    )
    parser.add_argument("--legacy-xlsx", help="Caminho do arquivo legado .xlsx")
    parser.add_argument("--credentials", help="Caminho do JSON da service account")
    parser.add_argument("--spreadsheet-id", help="ID da planilha de destino")
    parser.add_argument("--spreadsheet-name", help="Nome da planilha de destino (fallback)")
    parser.add_argument("--secrets-path", help="Caminho do secrets.toml")
    parser.add_argument(
        "--preview-xlsx",
        default=DEFAULT_PREVIEW_FILENAME,
        help=f"Arquivo de preview do dry-run (padrao: {DEFAULT_PREVIEW_FILENAME})",
    )
    parser.add_argument(
        "--skip-preview",
        action="store_true",
        help="Nao gerar arquivo de preview local",
    )
    parser.add_argument(
        "--commit",
        action="store_true",
        help="Confirma escrita no Google Sheets. Sem este flag, roda apenas simulacao.",
    )
    return parser.parse_args(argv)


def run(argv: list[str]) -> int:
    args = parse_args(argv)
    secrets_path = resolve_secrets_path(args.secrets_path)
    secrets = load_secrets(secrets_path)
    legacy_path = resolve_legacy_path(args.legacy_xlsx)

    credentials = get_credentials(args, secrets)
    client = gspread.authorize(credentials)

    spreadsheet_id, spreadsheet_name = resolve_target(args, secrets)
    spreadsheet = (
        client.open_by_key(spreadsheet_id)
        if spreadsheet_id
        else client.open(spreadsheet_name)
    )

    ws_frotas = spreadsheet.worksheet("frotas")
    ws_viagens = spreadsheet.worksheet("viagens")
    ws_despesas = spreadsheet.worksheet("despesas")

    frotas_header, frotas_rows_dict = get_header_and_rows(ws_frotas)
    viagens_header, viagens_rows_dict = get_header_and_rows(ws_viagens)
    despesas_header, despesas_rows_dict = get_header_and_rows(ws_despesas)

    ensure_required_columns("frotas", frotas_header)
    ensure_required_columns("viagens", viagens_header)
    ensure_required_columns("despesas", despesas_header)

    existing_frotas = {canonical_id(row.get("frota_id")) for row in frotas_rows_dict if row.get("frota_id")}

    existing_trip_sigs: set[tuple[str, str, str, float | None]] = set()
    max_viagem_id = 0
    for row in viagens_rows_dict:
        max_viagem_id = max(max_viagem_id, safe_int(row.get("viagem_id")))
        signature = (
            date_key(row.get("data_carregamento")),
            canonical_id(row.get("frota_id")),
            normalize_text(row.get("destino")).upper(),
            amount_key(row.get("frete_total")),
        )
        existing_trip_sigs.add(signature)

    existing_exp_sigs: set[tuple[str, str, float | None, str, str]] = set()
    max_despesa_id = 0
    for row in despesas_rows_dict:
        max_despesa_id = max(max_despesa_id, safe_int(row.get("despesa_id")))
        signature = (
            date_key(row.get("data")),
            canonical_id(row.get("frota_id")),
            amount_key(row.get("valor")),
            normalize_text(row.get("categoria")).upper(),
            normalize_text(row.get("obs")).upper(),
        )
        existing_exp_sigs.add(signature)

    next_viagem_id = max_viagem_id + 1
    next_despesa_id = max_despesa_id + 1

    legacy_workbook = openpyxl.load_workbook(legacy_path, data_only=True)

    append_frotas_payload: list[dict[str, Any]] = []
    append_viagens_payload: list[dict[str, Any]] = []
    append_despesas_payload: list[dict[str, Any]] = []
    summary_rows: list[list[Any]] = []

    for legacy_ws in legacy_workbook.worksheets:
        try:
            parsed = parse_legacy_sheet(legacy_ws)
        except ValueError as exc:
            print(f"[SKIP] {legacy_ws.title}: {exc}")
            continue

        sheet_name = parsed["sheet_name"]
        frota_id = parsed["frota_id"]
        motorista = parsed["motorista"]
        viagens = parsed["viagens"]
        abastecimentos = parsed["abastecimentos"]
        outros_gastos_total = parsed["outros_gastos_total"]

        if not frota_id:
            print(f"[SKIP] {sheet_name}: nao foi possivel identificar frota_id.")
            continue

        frota_id_str = str(frota_id)
        imported_viagens = 0
        imported_despesas = 0

        if frota_id_str not in existing_frotas:
            frota_nome = "TNorte" if frota_id == 1831 else f"Frota {frota_id}"
            append_frotas_payload.append(
                {
                    "frota_id": frota_id_str,
                    "frota_nome": frota_nome,
                    "motorista_nome": motorista,
                    "ativa": "TRUE",
                }
            )
            existing_frotas.add(frota_id_str)

        for viagem in viagens:
            signature = (
                fmt_date(viagem["data"]),
                frota_id_str,
                normalize_text(viagem["destino"]).upper(),
                round(float(viagem["frete_total"]), 2),
            )
            if signature in existing_trip_sigs:
                continue

            append_viagens_payload.append(
                {
                    "viagem_id": str(next_viagem_id),
                    "data_carregamento": fmt_date(viagem["data"]),
                    "data_finalizacao": fmt_date(viagem["data"]),
                    "dias_viagem": "",
                    "frota_id": frota_id_str,
                    "destino": normalize_text(viagem["destino"]),
                    "frete_total": round(float(viagem["frete_total"]), 2),
                    "status": "finalizada",
                    "mes_competencia": month_comp(viagem["data"]),
                    "valor_adiantamento": "",
                    "data_adiantamento": "",
                    "valor_quitacao": "",
                    "data_prevista_quitacao": "",
                    "data_quitacao": "",
                    "status_pagamento": "pendente",
                }
            )
            existing_trip_sigs.add(signature)
            next_viagem_id += 1
            imported_viagens += 1

        for abastecimento in abastecimentos:
            data_ref = abastecimento["data"] or (viagens[0]["data"] if viagens else None)
            obs = f"Importacao legado | {sheet_name} | abastecimento sem vinculo por viagem"
            signature = (
                fmt_date(data_ref),
                frota_id_str,
                round(float(abastecimento["valor"]), 2),
                "ABASTECIMENTO",
                obs.upper(),
            )
            if signature in existing_exp_sigs:
                continue

            append_despesas_payload.append(
                {
                    "despesa_id": str(next_despesa_id),
                    "data": fmt_date(data_ref),
                    "frota_id": frota_id_str,
                    "viagem_id": "",
                    "categoria": "Abastecimento",
                    "valor": round(float(abastecimento["valor"]), 2),
                    "tipo_pagamento": "a_vista",
                    "obs": obs,
                }
            )
            existing_exp_sigs.add(signature)
            next_despesa_id += 1
            imported_despesas += 1

        if outros_gastos_total and outros_gastos_total > 0:
            data_ref = max((v["data"] for v in viagens), default=None)
            obs = f"Importacao legado | {sheet_name} | total mensal agregado sem detalhamento"
            signature = (
                fmt_date(data_ref),
                frota_id_str,
                round(float(outros_gastos_total), 2),
                "OUTRAS DESPESAS",
                obs.upper(),
            )
            if signature not in existing_exp_sigs:
                append_despesas_payload.append(
                    {
                        "despesa_id": str(next_despesa_id),
                        "data": fmt_date(data_ref),
                        "frota_id": frota_id_str,
                        "viagem_id": "",
                        "categoria": "Outras despesas",
                        "valor": round(float(outros_gastos_total), 2),
                        "tipo_pagamento": "a_pagar",
                        "obs": obs,
                    }
                )
                existing_exp_sigs.add(signature)
                next_despesa_id += 1
                imported_despesas += 1

        summary_rows.append(
            [
                sheet_name,
                frota_id_str,
                motorista,
                len(viagens),
                len(abastecimentos),
                outros_gastos_total or 0,
                imported_viagens,
                imported_despesas,
            ]
        )

    append_frotas_rows = [row_from_payload(payload, frotas_header) for payload in append_frotas_payload]
    append_viagens_rows = [row_from_payload(payload, viagens_header) for payload in append_viagens_payload]
    append_despesas_rows = [row_from_payload(payload, despesas_header) for payload in append_despesas_payload]

    print("Resumo da migracao:")
    print(f"- Arquivo legado: {legacy_path}")
    print(f"- Frotas para adicionar: {len(append_frotas_rows)}")
    print(f"- Viagens para adicionar: {len(append_viagens_rows)}")
    print(f"- Despesas para adicionar: {len(append_despesas_rows)}")

    if not args.skip_preview:
        preview_path = Path(args.preview_xlsx).expanduser().resolve()
        build_preview(
            preview_path=preview_path,
            summary_rows=summary_rows,
            frotas_header=frotas_header,
            viagens_header=viagens_header,
            despesas_header=despesas_header,
            frotas_rows=append_frotas_rows,
            viagens_rows=append_viagens_rows,
            despesas_rows=append_despesas_rows,
        )
        print(f"- Preview gerado: {preview_path}")

    if not args.commit:
        print("Dry-run finalizado. Nenhuma linha foi gravada no Google Sheets.")
        print("Para gravar, rode novamente com --commit.")
        return 0

    if append_frotas_rows:
        ws_frotas.append_rows(append_frotas_rows, value_input_option="USER_ENTERED")
    if append_viagens_rows:
        ws_viagens.append_rows(append_viagens_rows, value_input_option="USER_ENTERED")
    if append_despesas_rows:
        ws_despesas.append_rows(append_despesas_rows, value_input_option="USER_ENTERED")

    print("Migracao concluida com gravacao.")
    return 0


def main() -> None:
    try:
        code = run(sys.argv[1:])
    except MigrationError as exc:
        print(f"[ERRO] {exc}")
        raise SystemExit(1) from exc
    except Exception as exc:  # pragma: no cover - fallback de runtime
        print(f"[ERRO] Falha inesperada: {exc}")
        raise SystemExit(1) from exc
    raise SystemExit(code)


if __name__ == "__main__":
    main()
